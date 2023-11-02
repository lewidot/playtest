"""File containing functions for loading test data."""
import json
from csv import reader
from pathlib import Path

from openpyxl import load_workbook


def _load_csv_data(path: Path) -> list[tuple]:
    """Load csv data for test parametrization."""
    with open(path) as f:
        # create csv reader object
        csv_reader = reader(f)

        # skip the header row
        next(csv_reader)

        # transform csv rows into a list of tuples
        return list(map(tuple, csv_reader))


def _load_json_data(path: Path) -> list[tuple]:
    """Load json data for test parametrization."""
    with Path.open(path, "r") as f:
        # load the json data
        data = json.load(f)

        # transform json values into a list of tuples
        return [tuple(d.values()) for d in data]


def _load_excel_data(path: Path) -> list[tuple]:
    """Load excel data for test parametrization.

    - Workbook must have single active sheet

    - Must have a header row with column headers

    - Header row columns must not be empty
    """
    # load workbook
    wb = load_workbook(path, read_only=True)

    # get the first worksheet
    sheet = wb.worksheets[0]

    # get the first row of the sheet (header row)
    header_row = next(sheet.rows)

    # get the index of the first header column without a value
    for idx, header in enumerate(header_row):
        if header.value == "":
            max_col_index = idx
            break
        else:
            max_col_index = sheet.max_column

    # create an iterator of rows, skipping header row (1 indexed)
    rows = sheet.iter_rows(
        min_row=2,
        max_row=sheet.max_row,
        max_col=max_col_index,
        values_only=True,
    )

    # return a list of tuples with data from each row
    return [row for row in rows if row[0] is not None]


def load_data(path: Path) -> list[tuple]:
    """Load data for test parametrization.

    Supports csv, json and excel file types

    """
    if not path.exists():
        err = f"{path} does not exist"
        raise FileNotFoundError(err)

    if path.suffix == ".csv":
        return _load_csv_data(path)

    elif path.suffix == ".json":
        return _load_json_data(path)

    elif path.suffix == ".xlsx":
        return _load_excel_data(path)

    else:
        err = f"file: {path} must be a csv, json or xlsx file"
        raise ValueError(err)
